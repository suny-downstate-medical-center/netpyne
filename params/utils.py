"""
utils.py 

Useful functions related to the parameters file, eg. create params file from excel table 

Contributors: salvadordura@gmail.com
"""
import os, sys
from neuron import h

def getSecName(sec, dirCellSecNames = {}):
	if '>.' in sec.name():
		fullSecName = sec.name().split('>.')[1] 
	elif '.' in sec.name():
		fullSecName = sec.name().split('.')[1]  
	else:
		fullSecName = sec.name()
	if '[' in fullSecName:  # if section is array element
		secNameTemp = fullSecName.split('[')[0]
		secIndex = int(fullSecName.split('[')[1].split(']')[0])
		secName = secNameTemp+'_'+str(secIndex) #(secNameTemp,secIndex)
	else:
		secName = fullSecName
		secIndex = -1
	if secName in dirCellSecNames:  # get sec names from python
		secName = dirCellSecNames[secName]
	return secName

def importCellParams(fileName, labels, values, key = None):
	params = {}
	if fileName.endswith('.py'):
		try:
	 		filePath,fileNameOnly = os.path.split(fileName)  # split path from filename
	  		if filePath not in sys.path:  # add to path if not there (need to import module)
	 			sys.path.insert(0, filePath)
			moduleName = fileNameOnly.split('.py')[0]  # remove .py to obtain module name
			exec('import '+ moduleName + ' as tempModule') in locals() # import module dynamically
			modulePointer = tempModule
			paramLabels = getattr(modulePointer, labels) # tuple with labels
			paramValues = getattr(modulePointer, values)  # variable with paramValues
			if key:  # if paramValues = dict
				paramValues = paramValues[key]
			params = dict(zip(paramLabels, paramValues))
			sys.path.remove(filePath)
		except:
			print "Error loading cell parameter values from " + fileName
	else:
		print "Trying to import izhi params from a file without the .py extension"
	return params


def mechVarList():
    msname = h.ref('')
    varList = {}
    for i, mechtype in enumerate(['mechs','pointps']):
        mt = h.MechanismType(i)  # either distributed mechs (0) or point process (1)
        varList[mechtype] = {}
        for j in xrange(int(mt.count())):
            mt.select(j)
            mt.selected(msname)
            ms = h.MechanismStandard(msname[0], 1) # PARAMETER (modifiable)
            varList[mechtype][msname[0]] = []
            propName = h.ref('')
            for var in xrange(int(ms.count())):
                k = ms.name(propName, var)
                varList[mechtype][msname[0]].append(propName[0])
    return varList

def importCell(fileName, cellName, cellArgs = {}):
	''' Import cell from HOC template or python file into framework format (dict of sections, with geom, topol, mechs, syns)'''
	if fileName.endswith('.hoc'):
		h.load_file(fileName)
		cell = getattr(h, cellName)(**cellArgs)  # arguments correspond to zloc, type and id -- remove in future (not used internally)
		secList = list(cell.allsec())
		dirCell = dir(cell)
	elif fileName.endswith('.py'):
 		filePath,fileNameOnly = os.path.split(fileName)  # split path from filename
  		if filePath not in sys.path:  # add to path if not there (need to import module)
 			sys.path.insert(0, filePath)
		moduleName = fileNameOnly.split('.py')[0]  # remove .py to obtain module name
		exec('import ' + moduleName + ' as tempModule') in globals(), locals() # import module dynamically
		modulePointer = tempModule
		cell = getattr(modulePointer, cellName)(**cellArgs)  # create cell and pass type as argument
		dirCell = dir(cell)

		if 'all_sec' in dirCell:
			secList = cell.all_sec
		elif 'sec' in dirCell:
			secList = [cell.sec]
		elif 'allsec' in dir(h):
			secList = [sec for sec in h.allsec()]
		elif 'soma' in dirCell:
			secList = [cell.soma]
		else:
			secList = []
		sys.path.remove(filePath)
	else:
		print "File name should be either .hoc or .py file"
		return

	# create dict with hname of each element in dir(cell)
	dirCellHnames = {}  
	for dirCellName in dirCell:
		try:
			dirCellHnames.update({cell.__dict__[dirCellName].hname(): dirCellName})
		except:
			pass
	# create dict with dir(cell) name corresponding to each hname 
	dirCellSecNames = {} 
	for sec in secList:
		dirCellSecNames.update({hname: name for hname,name in dirCellHnames.iteritems() if hname == sec.hname()})

	secDic = {}
	for sec in secList: 
		# create new section dict with name of section
		secName = getSecName(sec, dirCellSecNames)

		if len(secList) == 1:
			secName = 'soma' # if just one section rename to 'soma'
		secDic[secName] = {'geom': {}, 'topol': {}, 'mechs': {}, 'syns': {}}  # create dictionary to store sec info

		# store geometry properties
		standardGeomParams = ['L', 'nseg', 'diam', 'Ra', 'cm']
		secDir = dir(sec)
		for geomParam in standardGeomParams:
			#if geomParam in secDir:
			try:
				secDic[secName]['geom'][geomParam] = sec.__getattribute__(geomParam)
			except:
				pass

		# store 3d geometry
		numPoints = int(h.n3d())
		if numPoints: 
			points = []
			for ipoint in range(numPoints):
				x = h.x3d(ipoint)
				y = h.y3d(ipoint)
				z = h.z3d(ipoint)
				diam = h.diam3d(ipoint)
				points.append((x, y, z, diam))
			secDic[secName]['geom']['pt3d'] = points

		# store mechanisms
		varList = mechVarList()  # list of properties for all density mechanisms and point processes
		ignoreMechs = ['dist']  # dist only used during cell creation 
		mechDic = {}
		for mech in dir(sec(0.5)):  
			if h.ismembrane(mech) and mech not in ignoreMechs:  # check if membrane mechanism
				mechDic[mech] = {}  # create dic for mechanism properties
				varNames = [varName.replace('_'+mech, '') for varName in varList['mechs'][mech]]
				varVals = []
				for varName in varNames:
					try:
						varVals = [seg.__getattribute__(mech).__getattribute__(varName) for seg in sec]
						if len(set(varVals)) == 1:
							varVals = varVals[0] 
						mechDic[mech][varName] = varVals
					except: 
						pass
						#print 'Could not read variable %s from mechanism %s'%(varName,mech)

		secDic[secName]['mechs'] = mechDic

		# add synapses and point neurons
		# for now read fixed params, but need to find way to read only synapse params
		syns = {}
		pointps = {}
		for seg in sec:
			for ipoint,point in enumerate(seg.point_processes()):
				pptype = point.hname().split('[')[0]
				varNames = varList['pointps'][pptype]
				if any([s in pptype.lower() for s in ['syn', 'ampa', 'gaba', 'nmda', 'glu']]):
				#if 'syn' in pptype.lower(): # if syn in name of point process then assume synapse
					synName = pptype + '_' + str(len(syns))
					syns[synName] = {}
					syns[synName]['_type'] = pptype
					syns[synName]['_loc'] = seg.x
					for varName in varNames:
						try:
							syns[synName][varName] = point.__getattribute__(varName)
						except:
							print 'Could not read variable %s from synapse %s'%(varName,synName)
				
				else: # assume its a non-synapse point process
					pointpName = pptype + '_'+ str(len(pointps))
					pointps[pointpName] = {}
					pointps[pointpName]['_type'] = pptype
					pointps[pointpName]['_loc'] = seg.x
					for varName in varNames:
						try:
							pointps[pointpName][varName] = point.__getattribute__(varName)
							# special condition for Izhi model, to set vinit=vr
							# if varName == 'vr': secDic[secName]['vinit'] = point.__getattribute__(varName) 
						except:
							print 'Could not read %s variable from point process %s'%(varName,synName)

		if syns: secDic[secName]['syns'] = syns
		if pointps: secDic[secName]['pointps'] = pointps

		# store topology (keep at the end since h.SectionRef messes remaining loop)
		secRef = h.SectionRef(sec=sec)
		if secRef.has_parent():
			secDic[secName]['topol']['parentSec'] = getSecName(secRef.parent().sec, dirCellSecNames)
			secDic[secName]['topol']['parentX'] = h.parent_connection()
			secDic[secName]['topol']['childX'] = h.section_orientation()

	del(cell) # delete cell
	import gc; gc.collect()

	return secDic


def importConnFromExcel(fileName, sheetName):
	''' Import connectivity rules from Excel sheet'''
	import openpyxl as xl

	# set columns
	colPreTags = 0 # 'A'
	colPostTags = 1 # 'B'
	colConnFunc = 2 # 'C'
	colSyn = 3 # 'D'
	colProb = 5 # 'F'
	colWeight = 6 # 'G'
	colAnnot = 8 # 'I' 

	outFileName = fileName[:-5]+'_'+sheetName+'.py' # set output file name

	connText = """## Generated using importConnFromExcel() function in params/utils.py \n\nnetParams['connParams'] = [] \n\n"""
	
	# open excel file and sheet
	wb = xl.load_workbook(fileName)
	sheet = wb.get_sheet_by_name(sheetName)
	numRows = sheet.get_highest_row()

	with open(outFileName, 'w') as f:
		f.write(connText)  # write starting text
		for row in range(1,numRows+1):
			if sheet.cell(row=row, column=colProb).value:  # if not empty row
				print 'Creating conn rule for row ' + str(row)
				# read row values
				pre = sheet.cell(row=row, column=colPreTags).value
				post = sheet.cell(row=row, column=colPostTags).value
				func = sheet.cell(row=row, column=colConnFunc).value
				syn = sheet.cell(row=row, column=colSyn).value
				prob = sheet.cell(row=row, column=colProb).value
				weight = sheet.cell(row=row, column=colWeight).value

				# write preTags
				line = "netParams['connParams'].append({'preTags': {"
				for i,cond in enumerate(pre.split(';')):  # split into different conditions
					if i>0: line = line + ", "
					cond2 = cond.split('=')  # split into key and value
					line = line + "'" + cond2[0].replace(' ','') + "': " + cond2[1].replace(' ','')   # generate line
				line = line + "}" # end of preTags		

				# write postTags
				line = line + ",\n'postTags': {"
				for i,cond in enumerate(post.split(';')):  # split into different conditions
					if i>0: line = line + ", "
					cond2 = cond.split('=')  # split into key and value
					line = line + "'" + cond2[0].replace(' ','') + "': " + cond2[1].replace(' ','')   # generate line
				line = line + "}" # end of postTags			
				line = line + ",\n'connFunc': '" + func + "'"  # write connFunc
				line = line + ",\n'synReceptor': '" + syn + "'"  # write synReceptor
				line = line + ",\n'probability': " + str(prob)  # write prob
				line = line + ",\n'weight': " + str(weight)  # write prob
				line = line + "})"  # add closing brackets
				line = line + '\n\n' # new line after each conn rule
				f.write(line)  # write to file
				
