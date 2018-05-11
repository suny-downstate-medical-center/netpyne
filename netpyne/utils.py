"""
utils.py 

Useful functions

Contributors: salvador dura@gmail.com
"""
import os, sys
from numbers import Number
from neuron import h
import importlib

#h.load_file("stdrun.hoc") 

def getSecName (sec, dirCellSecNames = None):
    if dirCellSecNames is None: dirCellSecNames = {}

    if '>.' in sec.name():
        fullSecName = sec.name().split('>.')[1] 
    elif '.' in sec.name():
        fullSecName = sec.name().split('.')[1]  
    else:
        fullSecName = sec.name()
    if '[' in fullSecName:  # if section is array element
        secNameTemp = fullSecName.split('[')[0]
        secIndex = int(fullSecName.split('[')[1].split(']')[0])
        secName = secNameTemp+'_'+str(secIndex) 
    else:
        secName = fullSecName
        secIndex = -1
    if secName in dirCellSecNames:  # get sec names from python
        secName = dirCellSecNames[secName]
    return secName

def importCellParams (fileName, labels, values, key = None):
    params = {}
    if fileName.endswith('.py'):
        try:
            filePath,fileNameOnly = os.path.split(fileName)  # split path from filename
            if filePath not in sys.path:  # add to path if not there (need to import module)
                sys.path.insert(0, filePath)
                removeFilePath = True
            else:
                removeFilePath = False
            moduleName = fileNameOnly.split('.py')[0]  # remove .py to obtain module name
            tempModule = importlib.import_module(moduleName)
            modulePointer = tempModule
            paramLabels = getattr(modulePointer, labels) # tuple with labels
            paramValues = getattr(modulePointer, values)  # variable with paramValues
            if key:  # if paramValues = dict
                paramValues = paramValues[key]
            params = dict(zip(paramLabels, paramValues))
            if removeFilePath: sys.path.remove(filePath)
        except:
            print "Error loading cell parameter values from " + fileName
    else:
        print "Trying to import izhi params from a file without the .py extension"
    return params


def mechVarList ():
    msname = h.ref('')
    varList = {}
    for i, mechtype in enumerate(['mechs','pointps']):
        mt = h.MechanismType(i)  # either distributed mechs (0) or point process (1)
        varList[mechtype] = {}
        for j in xrange(int(mt.count())):
            mt.select(j)
            mt.selected(msname)
            ms = h.MechanismStandard(msname[0], 1) # PARAMETER (modifiable)
            varList[mechtype][msname[0]] = []
            propName = h.ref('')
            for var in xrange(int(ms.count())):
                k = ms.name(propName, var)
                varList[mechtype][msname[0]].append(propName[0])
    return varList


# def getAllGlobals (origGlob={}):
#     exclude = ['nseg', 'nrn_shape_changed_', 'Ra', 'L', 'cm', 'rallbranch', 'tstop']  # don't check these because crashes
#     glob = {}
#     # compare with original
#     for k in [x for x in dir(h) if x not in exclude]:
#         try:
#             v = getattr(h,k)
#             if isinstance(v, Number):     # store float and int globals 
#                 if k not in origGlob or origGlob[k] != v:
#                     glob[k] = v
#         except:
#             pass
#     return glob


def getGlobals (mechNames, origGlob={}):
    includeGlobs = ['celsius', 'v_init', 'clamp_resist']
    endings = tuple(['_' + name for name in mechNames])
    glob = {}
    # compare with original
    for k in dir(h):
        if k.endswith(endings) or k in includeGlobs:
            try:
                v = float(h.__getattribute__(k))
                if k not in origGlob or origGlob[k] != v:
                    glob[k] = float(v)
            except:
                pass
    return glob


def setGlobals (glob):
    for k,v in glob.iteritems():
        try:
            setattr(h, k, v)
        except:
            pass

    # # remove vars are not in glob ?
    # for k in [x for x in dir(h) if x not in exclude]:
    #     if k not in glob:
    #         try:
    #             setattr(h, k, None) 
    #         except:
    #             print k


def _equal_dicts (d1, d2, ignore_keys):
    ignored = set(ignore_keys)
    for k1, v1 in d1.iteritems():
        if k1 not in ignored and (k1 not in d2 or d2[k1] != v1):
            return False
    for k2, v2 in d2.iteritems():
        if k2 not in ignored and k2 not in d1:
            return False
    return True


def _delete_module(modname):
    from sys import modules
    try:
        thismod = modules[modname]
        del modules[modname]
    except KeyError:
        pass

    for mod in modules.values():
        try:
            delattr(mod, modname)
        except:
            pass

def importCell (fileName, cellName, cellArgs = None, cellInstance = False):
    h.initnrn()
    varList = mechVarList()  # list of properties for all density mechanisms and point processes
    origGlob = getGlobals(varList['mechs'].keys()+varList['pointps'].keys())
    origGlob['v_init'] = -65  # add by hand since won't be set unless load h.load_file('stdrun')

    if cellArgs is None: cellArgs = [] # Define as empty list if not otherwise defined

    ''' Import cell from HOC template or python file into framework format (dict of sections, with geom, topol, mechs, syns)'''
    if fileName.endswith('.hoc') or fileName.endswith('.tem'):
        h.load_file(fileName)
        if not cellInstance:
            if isinstance(cellArgs, dict):
                cell = getattr(h, cellName)(**cellArgs)  # create cell using template, passing dict with args
            else:
                cell = getattr(h, cellName)(*cellArgs) # create cell using template, passing list with args
        else:
            try:
                cell = getattr(h, cellName)
            except:
                cell = None
    elif fileName.endswith('.py'):
        filePath,fileNameOnly = os.path.split(fileName)  # split path from filename
        if filePath not in sys.path:  # add to path if not there (need to import module)
            sys.path.insert(0, filePath)
            removeFilePath = True
        else:
            removeFilePath = False
        moduleName = fileNameOnly.split('.py')[0]  # remove .py to obtain module name
        tempModule = importlib.import_module(moduleName)
        modulePointer = tempModule
        if isinstance(cellArgs, dict):
            cell = getattr(modulePointer, cellName)(**cellArgs) # create cell using template, passing dict with args
        else:
            cell = getattr(modulePointer, cellName)(*cellArgs)  # create cell using template, passing list with args
        if removeFilePath: sys.path.remove(filePath)
    else:
        print "File name should be either .hoc or .py file"
        return

    secDic, secListDic, synMechs, globs = getCellParams(cell, varList, origGlob)
    
    if fileName.endswith('.py'):
        _delete_module(moduleName)
        _delete_module('tempModule')
        del modulePointer
    elif fileName.endswith('.hoc'):
        for sec in h.allsec():
            try:
                sec.push()
                h.delete_section(sec=sec)
                h.pop_section()
            except:
                pass
    h.initnrn()

    setGlobals(origGlob)  # restore original globals

    return secDic, secListDic, synMechs, globs


def importCellsFromNet (netParams, fileName, labelList, condsList, cellNamesList, importSynMechs):
    h.initnrn()
    
    ''' Import cell from HOC template or python file into framework format (dict of sections, with geom, topol, mechs, syns)'''
    if fileName.endswith('.hoc') or fileName.endswith('.tem'):
        print 'Importing from .hoc network not yet supported'
        return
        # h.load_file(fileName)
        # for cellName in cellNames:
        #     cell = getattr(h, cellName) # create cell using template, passing dict with args
        #     secDic, secListDic, synMechs = getCellParams(cell)

    elif fileName.endswith('.py'):
        origDir = os.getcwd()
        filePath,fileNameOnly = os.path.split(fileName)  # split path from filename
        if filePath not in sys.path:  # add to path if not there (need to import module)
            sys.path.insert(0, filePath)
            removeFilePath = True
        else:
            removeFilePath = False
        moduleName = fileNameOnly.split('.py')[0]  # remove .py to obtain module name
        os.chdir(filePath)
        print '\nRunning network in %s to import cells into NetPyNE ...\n'%(fileName)
        from neuron import load_mechanisms
        load_mechanisms(filePath)
        tempModule = importlib.import_module(moduleName)
        modulePointer = tempModule
        if removeFilePath: sys.path.remove(filePath)
    else:
        print "File name should be either .hoc or .py file"
        return

    for label, conds, cellName in zip(labelList, condsList, cellNamesList):
        print '\nImporting %s from %s ...'%(cellName, fileName)
        exec('cell = tempModule' + '.' + cellName)
        #cell = getattr(modulePointer, cellName) # get cell object
        varList = mechVarList()
        origGlob = getGlobals(varList['mechs'].keys()+varList['pointps'].keys())
        secs, secLists, synMechs = getCellParams(cell, varList, origGlob)
        cellRule = {'conds': conds, 'secs': secs, 'secLists': secLists}
        netParams.addCellParams(label, cellRule)
        if importSynMechs:
            for synMech in synMechs: netParams.addSynMechParams(synMech.pop('label'), synMech)


def getCellParams(cell, varList={}, origGlob={}):
    dirCell = dir(cell)

    if 'all_sec' in dirCell:
        secs = cell.all_sec
    elif 'sec' in dirCell:
        secs = [cell.sec]
    elif 'allsec' in dir(h):
        secs = [sec for sec in h.allsec()]
    elif 'soma' in dirCell:
        secs = [cell.soma]
    else:
        secs = []

    # create dict with hname of each element in dir(cell)
    dirCellHnames = {}  
    for dirCellName in dirCell:
        try:
            dirCellHnames.update({getattr(cell, dirCellName).hname(): dirCellName})
        except:
            pass
    # create dict with dir(cell) name corresponding to each hname 
    dirCellSecNames = {} 
    for sec in secs:
        dirCellSecNames.update({hname: name for hname,name in dirCellHnames.iteritems() if hname == sec.hname()})

    secDic = {}
    synMechs = []
    for sec in secs: 
        # create new section dict with name of section
        secName = getSecName(sec, dirCellSecNames)

        # if len(secs) == 1: secName = 'soma' # if just one section rename to 'soma' -- REMOVED, doesn't always apply
        secDic[secName] = {'geom': {}, 'topol': {}, 'mechs': {}}  # create dictionary to store sec info

        # store geometry properties
        standardGeomParams = ['L', 'nseg', 'diam', 'Ra', 'cm']
        secDir = dir(sec)
        for geomParam in standardGeomParams:
            #if geomParam in secDir:
            try:
                secDic[secName]['geom'][geomParam] = sec.__getattribute__(geomParam)
            except:
                pass

        # store 3d geometry
        sec.push()  # access current section so ismembrane() works
        numPoints = int(h.n3d())
        if numPoints: 
            points = []
            for ipoint in range(numPoints):
                x = h.x3d(ipoint)
                y = h.y3d(ipoint)
                z = h.z3d(ipoint)
                diam = h.diam3d(ipoint)
                points.append((x, y, z, diam))
            secDic[secName]['geom']['pt3d'] = points

        # store mechanisms
        #varList = mechVarList()  # list of properties for all density mechanisms and point processes
        ignoreMechs = ['dist']  # dist only used during cell creation 
        ignoreVars = []  # 
        mechDic = {}
        ionDic = {}
        
        for mech in dir(sec(0.5)): 
            if h.ismembrane(mech) and mech not in ignoreMechs:  # check if membrane mechanism
                if not mech.endswith('_ion'):  # exclude ions
                    mechDic[mech] = {}  # create dic for mechanism properties
                    varNames = [varName.replace('_'+mech, '') for varName in varList['mechs'][mech]]
                    varVals = []
                    for varName in varNames:
                        if varName not in ignoreVars:
                            try:
                                varVals = [seg.__getattribute__(mech).__getattribute__(varName) for seg in sec]
                                if len(set(varVals)) == 1:
                                    varVals = varVals[0] 
                                mechDic[mech][varName] = varVals
                            except: 
                                pass
                                #print 'Could not read variable %s from mechanism %s'%(varName,mech)

                # store ions
                elif mech.endswith('_ion'):
                    ionName = mech.split('_ion')[0]
                    varNames = [varName.replace('_'+mech, '').replace(ionName,'') for varName in varList['mechs'][mech]]
                    varNames.append('e')
                    varVals = []
                    ionDic[ionName] = {}  # create dic for mechanism properties
                    for varName in varNames:
                        varNameSplit = varName
                        if varName not in ignoreVars:
                            try:
                                if varNameSplit in ['i','o']: # var name after ion name (eg. 'nai', 'nao')
                                    varVals = [seg.__getattribute__(ionName+varNameSplit) for seg in sec]
                                else: # var name before ion name (eg. 'ena')
                                    varVals = [seg.__getattribute__(varNameSplit+ionName) for seg in sec]
                                if len(set(varVals)) == 1:
                                    varVals = varVals[0] 
                                ionDic[ionName][varNameSplit] = varVals
                            except: 
                                pass
                                #print 'Could not read variable %s from mechanism %s'%(varName,mech)                    


        secDic[secName]['mechs'] = mechDic
        if len(ionDic)>0: 
            secDic[secName]['ions'] = ionDic

        # add synapses and point neurons
        # for now read fixed params, but need to find way to read only synapse params
        pointps = {}
        for seg in sec:
            for ipoint,point in enumerate(seg.point_processes()):
                pointpMod = point.hname().split('[')[0]
                varNames = varList['pointps'][pointpMod]
                if any([s in pointpMod.lower() for s in ['syn', 'ampa', 'gaba', 'nmda', 'glu']]):
                #if 'synMech' in pptype.lower(): # if syn in name of point process then assume synapse
                    synMech = {}
                    synMech['label'] = pointpMod + '_' + str(len(synMechs))
                    synMech['mod'] = pointpMod
                    #synMech['loc'] = seg.x
                    for varName in varNames:
                        try:
                            synMech[varName] = point.__getattribute__(varName)
                        except:
                            print 'Could not read variable %s from synapse %s'%(varName,synMech['label'])

                    if not any([_equal_dicts(synMech, synMech2, ignore_keys=['label']) for synMech2 in synMechs]):
                        synMechs.append(synMech)
                
                else: # assume its a non-synapse point process
                    pointpName = pointpMod + '_'+ str(len(pointps))
                    pointps[pointpName] = {}
                    pointps[pointpName]['mod'] = pointpMod
                    pointps[pointpName]['loc'] = seg.x
                    for varName in varNames:
                        try:
                            pointps[pointpName][varName] = point.__getattribute__(varName)
                            # special condition for Izhi model, to set vinit=vr
                            # if varName == 'vr': secDic[secName]['vinit'] = point.__getattribute__(varName) 
                        except:
                            print 'Could not read %s variable from point process %s'%(varName,pointpName)

        if pointps: secDic[secName]['pointps'] = pointps

        # store topology (keep at the end since h.SectionRef messes remaining loop)
        secRef = h.SectionRef(sec=sec)
        if secRef.has_parent():
            secDic[secName]['topol']['parentSec'] = getSecName(secRef.parent().sec, dirCellSecNames)
            secDic[secName]['topol']['parentX'] = h.parent_connection()
            secDic[secName]['topol']['childX'] = h.section_orientation()

        h.pop_section()  # to prevent section stack overflow
        
    # store section lists
    secLists = h.List('SectionList')
    if int(secLists.count()): 
        secListDic = {}
        for i in xrange(int(secLists.count())):  # loop over section lists
            hname = secLists.o(i).hname()
            if hname in dirCellHnames:  # use python variable name
                secListName = dirCellHnames[hname]
            else:
                secListName = hname
            secListDic[secListName] = [getSecName(sec, dirCellSecNames) for sec in secLists.o(i)]
    else:
        secListDic = {}

    # globals
    globs = getGlobals(varList['mechs'].keys()+varList['pointps'].keys(), origGlob=origGlob)
    if 'v_init' in globs:  # set v_init for each section (allows for cells with differnet vinits)
        for sec in secDic.values(): sec['vinit'] = globs['v_init']  

    # clean 
    cell = None
    for i in range(len(secs)):
        tmp=secs.pop()
        del tmp

    import gc; gc.collect()


    return secDic, secListDic, synMechs, globs



def importConnFromExcel (fileName, sheetName):
    ''' Import connectivity rules from Excel sheet'''
    import openpyxl as xl

    # set columns
    colPreTags = 0 # 'A'
    colPostTags = 1 # 'B'
    colConnFunc = 2 # 'C'
    colSyn = 3 # 'D'
    colProb = 5 # 'F'
    colWeight = 6 # 'G'
    colAnnot = 8 # 'I' 

    outFileName = fileName[:-5]+'_'+sheetName+'.py' # set output file name

    connText = """## Generated using importConnFromExcel() function in params/utils.py \n\nnetParams['connParams'] = [] \n\n"""
    
    # open excel file and sheet
    wb = xl.load_workbook(fileName)
    sheet = wb.get_sheet_by_name(sheetName)
    numRows = sheet.get_highest_row()

    with open(outFileName, 'w') as f:
        f.write(connText)  # write starting text
        for row in range(1,numRows+1):
            if sheet.cell(row=row, column=colProb).value:  # if not empty row
                print 'Creating conn rule for row ' + str(row)
                # read row values
                pre = sheet.cell(row=row, column=colPreTags).value
                post = sheet.cell(row=row, column=colPostTags).value
                func = sheet.cell(row=row, column=colConnFunc).value
                syn = sheet.cell(row=row, column=colSyn).value
                prob = sheet.cell(row=row, column=colProb).value
                weight = sheet.cell(row=row, column=colWeight).value

                # write preTags
                line = "netParams['connParams'].append({'preConds': {"
                for i,cond in enumerate(pre.split(';')):  # split into different conditions
                    if i>0: line = line + ", "
                    cond2 = cond.split('=')  # split into key and value
                    line = line + "'" + cond2[0].replace(' ','') + "': " + cond2[1].replace(' ','')   # generate line
                line = line + "}" # end of preTags      

                # write postTags
                line = line + ",\n'postConds': {"
                for i,cond in enumerate(post.split(';')):  # split into different conditions
                    if i>0: line = line + ", "
                    cond2 = cond.split('=')  # split into key and value
                    line = line + "'" + cond2[0].replace(' ','') + "': " + cond2[1].replace(' ','')   # generate line
                line = line + "}" # end of postTags         
                line = line + ",\n'connFunc': '" + func + "'"  # write connFunc
                line = line + ",\n'synMech': '" + syn + "'"  # write synReceptor
                line = line + ",\n'probability': " + str(prob)  # write prob
                line = line + ",\n'weight': " + str(weight)  # write prob
                line = line + "})"  # add closing brackets
                line = line + '\n\n' # new line after each conn rule
                f.write(line)  # write to file
                
        
