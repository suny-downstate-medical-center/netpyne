"""
conversion/excel.py 

Functions to import from Excel
"""

from __future__ import print_function
from __future__ import unicode_literals
from __future__ import division
from __future__ import absolute_import


from builtins import open
from builtins import range
from builtins import str
from future import standard_library
standard_library.install_aliases()
def importConnFromExcel (fileName, sheetName):
    """
    Import connectivity rules from Excel sheet
    """
    
    import openpyxl as xl

    # set columns
    colPreTags = 0 # 'A'
    colPostTags = 1 # 'B'
    colConnFunc = 2 # 'C'
    colSyn = 3 # 'D'
    colProb = 5 # 'F'
    colWeight = 6 # 'G'
    colAnnot = 8 # 'I' 

    outFileName = fileName[:-5]+'_'+sheetName+'.py' # set output file name

    connText = """## Generated using importConnFromExcel() function in params/utils.py \n\nnetParams['connParams'] = [] \n\n"""
    
    # open excel file and sheet
    wb = xl.load_workbook(fileName)
    sheet = wb.get_sheet_by_name(sheetName)
    numRows = sheet.get_highest_row()

    with open(outFileName, 'w') as f:
        f.write(connText)  # write starting text
        for row in range(1,numRows+1):
            if sheet.cell(row=row, column=colProb).value:  # if not empty row
                print('Creating conn rule for row ' + str(row))
                # read row values
                pre = sheet.cell(row=row, column=colPreTags).value
                post = sheet.cell(row=row, column=colPostTags).value
                func = sheet.cell(row=row, column=colConnFunc).value
                syn = sheet.cell(row=row, column=colSyn).value
                prob = sheet.cell(row=row, column=colProb).value
                weight = sheet.cell(row=row, column=colWeight).value

                # write preTags
                line = "netParams['connParams'].append({'preConds': {"
                for i,cond in enumerate(pre.split(';')):  # split into different conditions
                    if i>0: line = line + ", "
                    cond2 = cond.split('=')  # split into key and value
                    line = line + "'" + cond2[0].replace(' ','') + "': " + cond2[1].replace(' ','')   # generate line
                line = line + "}" # end of preTags      

                # write postTags
                line = line + ",\n'postConds': {"
                for i,cond in enumerate(post.split(';')):  # split into different conditions
                    if i>0: line = line + ", "
                    cond2 = cond.split('=')  # split into key and value
                    line = line + "'" + cond2[0].replace(' ','') + "': " + cond2[1].replace(' ','')   # generate line
                line = line + "}" # end of postTags         
                line = line + ",\n'connFunc': '" + func + "'"  # write connFunc
                line = line + ",\n'synMech': '" + syn + "'"  # write synReceptor
                line = line + ",\n'probability': " + str(prob)  # write prob
                line = line + ",\n'weight': " + str(weight)  # write prob
                line = line + "})"  # add closing brackets
                line = line + '\n\n' # new line after each conn rule
                f.write(line)  # write to file
                